from openpyxl import Workbook
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.drawing.image import Image
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, IconSetRule
import pandas as pd

font_heading = Font(size=14, bold=True)

table_style = TableStyleInfo(
    name="TableStyleMedium2",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False,
)

# "#63be7b", "#ffffff", "#f8696b"
logfc_colorscale = ColorScaleRule(
    start_type="num",
    start_value=-6,
    start_color="63be7b",
    mid_type="num",
    mid_value=0,
    mid_color="ffffff",
    end_type="num",
    end_value=6,
    end_color="f8696b",
)




def make_range(cell, shape, header=False):
    rows, cols = shape
    col_start_char = cell[0]
    col_start = column_index_from_string(col_start_char)
    row_start = int(cell[1])

    col_end = get_column_letter(col_start + cols - 1)
    row_end = row_start + rows - 1 + header

    range = f"{col_start_char}{row_start}:{col_end}{row_end}"
    return range


def make_workbook():
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def add_worksheet(wb, name, heading=None, style=font_heading):
    ws = wb.create_sheet(name)
    if heading:
        ws["A1"] = heading
        ws["A1"].font = font_heading

    return ws


def add_table(
    ws, df, name, start_row=None, start_col=None, style=table_style, col_type=None
):
    if start_row is None:
        start_row = ws._current_row + 1
    if start_col is None:
        start_col = "A"
    cell = f"{start_col}{start_row}"

    ws._current_row = start_row - 1

    for idx, el in enumerate(df.columns):
        row = start_row
        col = column_index_from_string(start_col) + idx
        ws.cell(row=row, column=col, value=el)

    for df_row_idx, df_row in df.iterrows():
        for df_col_idx, value in enumerate(df_row.values):
            row = start_row + 1 + df_row_idx
            col = column_index_from_string(start_col) + df_col_idx
            c = ws.cell(row=row, column=col, value=value)
            if col_type:
                t = None
                if isinstance(col_type, str) or isinstance(col_type, list):
                    t = col_type[df_col_idx]
                elif isinstance(col_type, dict):
                    if df_col_idx in col_type.keys():
                        t = col_type[df_col_idx]
                    elif df.columns[df_col_idx] in col_type.keys():
                        t = col_type[df.columns[df_col_idx]]

                if t:
                    if t == "n":
                        c.number_format = "0.00"
                    elif t == "i":
                        c.number_format = "#,##0"
                    elif t == "s":
                        c.number_format = "@"
                    elif t == "e":
                        c.number_format = "0.00E+00"
                    else:
                        c.number_format = t

    tab = Table(displayName=name, ref=make_range(cell, df.shape, True))

    tab.tableStyleInfo = style

    ws.add_table(tab)


def set_column_width(ws, width):
    for i in range(ws.max_column):
        ws.column_dimensions[get_column_letter(i + 1)].width = width


def add_image(ws, anchor_cell, file, width=None, height=None):
    img = Image(file)
    width_original = img.width
    height_original = img.height

    if width and height:
        img.width = width
        img.height = height

    if width and not height:
        img.width = width
        img.height = height_original * img.width / width_original

    if not width and height:
        img.height = height
        img.width = width_original * img.height / height_original

    ws.add_image(img, anchor_cell)
